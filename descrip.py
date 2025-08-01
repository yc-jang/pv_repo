import re
from typing import Optional, Tuple, List, Union
import numpy as np
import pandas as pd

# ============== 유틸 ==============

def _as_str(x):
    if x is None:
        return ""
    s = str(x)
    return "" if s.strip().lower() in {"", "none", "nan"} else s

def _is_numeric_like(x) -> bool:
    if x is None:
        return False
    if isinstance(x, (int, float, np.number)):
        return True
    try:
        float(str(x).replace(",", ""))
        return True
    except Exception:
        return False

def _range_intersection(a, b):
    # a=(r1,c1,r2,c2), b=(r1,c1,r2,c2) 1-based inclusive
    r1 = max(a[0], b[0]); c1 = max(a[1], b[1])
    r2 = min(a[2], b[2]); c2 = min(a[3], b[3])
    if r1 <= r2 and c1 <= c2:
        return (r1, c1, r2, c2)
    return None

# ============== 경로 A: openpyxl로 병합 복원 읽기 ==============

def read_excel_with_merges(
    path: str,
    sheet: Union[int, str] = 0,
    area: Optional[str] = None,  # "A1:E200" 같은 범위를 지정 가능
    data_only: bool = True
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    병합 셀을 실제 값으로 펼친 그리드(DataFrame)와, 병합/채움으로 채워진 셀 마스크(bool)를 반환합니다.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.cell import coordinate_to_tuple
    from openpyxl.utils import range_boundaries

    wb = load_workbook(path, data_only=data_only)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]

    if area is None:
        min_row, max_row = ws.min_row, ws.max_row
        min_col, max_col = ws.min_column, ws.max_column
    else:
        c1, r1, c2, r2 = range_boundaries(area)  # note: returns (min_col,min_row,max_col,max_row)
        min_row, max_row = r1, r2
        min_col, max_col = c1, c2

    nrows = max_row - min_row + 1
    ncols = max_col - min_col + 1

    grid = [[ws.cell(row=r, column=c).value for c in range(min_col, max_col+1)] for r in range(min_row, max_row+1)]
    filled_by_merge = [[False]*ncols for _ in range(nrows)]

    # 병합 범위 전개
    for merged in ws.merged_cells.ranges:
        # openpyxl's merged: like "A1:C2"
        from openpyxl.utils import range_boundaries
        mc1, mr1, mc2, mr2 = range_boundaries(str(merged))
        inter = _range_intersection((min_row, min_col, max_row, max_col), (mr1, mc1, mr2, mc2))
        if inter is None:
            continue
        r1, c1, r2, c2 = inter
        v = ws.cell(row=mr1, column=mc1).value  # top-left of original merge
        for rr in range(r1, r2+1):
            for cc in range(c1, c2+1):
                grid[rr-min_row][cc-min_col] = v
                filled_by_merge[rr-min_row][cc-min_col] = True

    df = pd.DataFrame(grid)
    mask = pd.DataFrame(filled_by_merge, index=df.index, columns=df.columns)
    return df, mask

# ============== 경로 B: 클립보드에서 읽기 (헤더/스텁만 제한 보정) ==============

def read_clipboard_grid() -> pd.DataFrame:
    """
    엑셀에서 범위를 복사해 온 상태에서 호출.
    """
    df = pd.read_clipboard(sep="\t", header=None)
    # Excel에서 복사하면 수식은 값으로 붙고, 병합은 사라져 공칸으로 들어옵니다.
    return df

# ============== 헤더/스텁 자동 추정 ==============

def infer_header_depth(df: pd.DataFrame, probe_rows: int = 10, focus_cols: Optional[List[int]] = None) -> int:
    """
    상단 몇 행이 헤더인지 추정.
    - 규칙(휴리스틱):
      1) 위에서부터 훑으면서 숫자형 비율이 낮고 문자열 비율이 높은 행은 '헤더'로 간주
      2) 특정 행부터 숫자형 비율이 뚜렷이 높아지면 그 직전까지를 헤더로 확정
    - focus_cols: (예: [0,1,2,3,4]) A~E 열만 보고 결정하고 싶을 때 사용
    """
    n = min(probe_rows, len(df))
    cols = focus_cols if focus_cols is not None else list(range(min(5, df.shape[1])))
    depth = 0
    for r in range(n):
        row = df.iloc[r, cols]
        vals = list(row)
        if len(vals) == 0:
            break
        num_ratio = np.mean([_is_numeric_like(v) for v in vals])
        str_ratio = np.mean([(isinstance(v, str) and _as_str(v) != "") for v in vals])
        # 헤더는 보통 문자열 위주, 숫자 비율 낮음
        if str_ratio >= 0.3 and num_ratio < 0.5:
            depth += 1
        else:
            break
    return max(depth, 1)  # 최소 1행 헤더로

def infer_stub_width(df: pd.DataFrame, max_check: int = 8, look_rows: int = 30) -> int:
    """
    좌측 스텁(행 헤더) 폭 추정: 왼쪽에서부터 숫자형 비율이 낮은 열을 카운트.
    """
    rows = min(look_rows, df.shape[0])
    width = 0
    for c in range(min(max_check, df.shape[1])):
        col = df.iloc[:rows, c]
        num_ratio = np.mean([_is_numeric_like(v) for v in col])
        # 스텁은 보통 텍스트가 많고 숫자형 비율이 낮다.
        if num_ratio <= 0.4:
            width += 1
        else:
            break
    return width

# ============== 멀티헤더/스텁 보정 유틸 ==============

def fill_horizontal_in_rows(df: pd.DataFrame, row_idx: List[int]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    지정된 행들에 한해, 좌->우 수평 ffill (클립보드 경로에서 헤더 복원에 사용)
    반환: (채운 DF, 채움마스크)
    """
    out = df.copy()
    mask = pd.DataFrame(False, index=df.index, columns=df.columns)
    for r in row_idx:
        vals = out.iloc[r].astype(object).tolist()
        last = None
        for c in range(len(vals)):
            if _as_str(vals[c]) == "":
                if last is not None:
                    vals[c] = last
                    mask.iat[r, c] = True
            else:
                last = vals[c]
        out.iloc[r] = vals
    return out, mask

def fill_vertical_in_cols(df: pd.DataFrame, col_idx: List[int], start_row: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    지정된 열들에 한해, start_row부터 아래로만 세로 ffill (스텁 복원용)
    반환: (채운 DF, 채움마스크)
    """
    out = df.copy()
    mask = pd.DataFrame(False, index=df.index, columns=df.columns)
    for c in col_idx:
        last = None
        for r in range(start_row, df.shape[0]):
            v = out.iat[r, c]
            if _as_str(v) == "":
                if last is not None:
                    out.iat[r, c] = last
                    mask.iat[r, c] = True
            else:
                last = v
    return out, mask

def make_multiindex_columns_from_header_rows(df: pd.DataFrame, header_depth: int) -> pd.MultiIndex:
    """
    상단 header_depth 행을 이용해 MultiIndex 컬럼 구성.
    (행 단위로 수평 ffill을 선행했다고 가정하거나, openpyxl 경로에서는 병합 복원이 되어있음)
    """
    header = df.iloc[:header_depth].copy()
    header = header.applymap(_as_str)

    # 수평 ffill로 남은 빈칸 보정(오염 최소화를 위해 header 영역에만)
    for r in range(header_depth):
        header.iloc[r] = header.iloc[r].replace("", np.nan).ffill(axis=0)

    # 모두 빈 레벨 제거
    keep_levels = []
    arrays = []
    for r in range(header_depth):
        arr = header.iloc[r].tolist()
        if any(_as_str(x) != "" for x in arr):
            arrays.append([_as_str(x) for x in arr])
            keep_levels.append(r)

    if not arrays:
        # 전부 비어있다면 단일 레벨로 대체
        arrays = [ [f"col_{i}" for i in range(df.shape[1])] ]

    names = [f"level_{i+1}" for i in range(len(arrays))]
    return pd.MultiIndex.from_arrays(arrays, names=names)

# ============== 파이프라인 진입점 ==============

def normalize_excel_like_table(
    source: str = None,
    use_openpyxl: bool = True,
    sheet: Union[int, str] = 0,
    area: Optional[str] = None,
    override_header_depth: Optional[int] = None,
    override_stub_width: Optional[int] = None,
    focus_cols_for_header: Optional[List[int]] = None
) -> Tuple[pd.DataFrame, dict]:
    """
    복잡 병합/멀티헤더 테이블을 정규화:
      - columns: MultiIndex
      - index: 기본 RangeIndex (원하면 이후에 스텁 열로 MultiIndex index 구성 가능)
    반환: (정규화 DF, 메타정보 dict)
    """
    if use_openpyxl:
        raw, merge_mask = read_excel_with_merges(source, sheet=sheet, area=area)
        src = "openpyxl"
    else:
        raw = read_clipboard_grid()
        merge_mask = pd.DataFrame(False, index=raw.index, columns=raw.columns)
        src = "clipboard"

    # 헤더 깊이/스텁 폭 추정 (A~E를 우선 보겠다면 focus_cols_for_header = [0,1,2,3,4])
    header_depth = override_header_depth or infer_header_depth(raw, focus_cols=focus_cols_for_header)
    stub_width = override_stub_width or infer_stub_width(raw)

    filled_mask = merge_mask.copy()

    # 클립보드 경로인 경우에만 '스코프 제한 보정'
    if src == "clipboard":
        # 1) 헤더 영역 수평 보정
        raw, m1 = fill_horizontal_in_rows(raw, row_idx=list(range(header_depth)))
        filled_mask = (filled_mask | m1)

        # 2) 좌측 스텁 영역 세로 보정 (데이터 시작 행부터)
        raw, m2 = fill_vertical_in_cols(raw, col_idx=list(range(stub_width)), start_row=header_depth)
        filled_mask = (filled_mask | m2)

    # 컬럼 MultiIndex 구성
    cols = make_multiindex_columns_from_header_rows(raw, header_depth=header_depth)
    data = raw.iloc[header_depth:].reset_index(drop=True)
    data.columns = cols

    meta = dict(
        source=src,
        header_depth=header_depth,
        stub_width=stub_width,
        filled_mask=filled_mask  # True인 곳은 병합복원/보정으로 채워진 셀
    )
    return data, meta

# ============== (선택) 좌측 스텁으로 행 MultiIndex 구성 ==============

def attach_row_multiindex_from_stub(
    df: pd.DataFrame,
    raw_source: pd.DataFrame,
    header_depth: int,
    stub_width: int
) -> pd.DataFrame:
    """
    상단에서 정규화한 df에 대해, 좌측 stub_width 열을 행 MultiIndex로 사용하고 본문 데이터만 남기는 도우미.
    """
    if stub_width <= 0:
        return df
    stub = raw_source.iloc[header_depth:, :stub_width].copy()
    stub = stub.applymap(_as_str)
    # 스텁 내부 세로 ffill은 이미 수행되었다고 가정(클립보드) 또는 병합복원(openpyxl)
    idx = pd.MultiIndex.from_frame(stub)
    body = df.copy()
    body.index = idx
    return body
